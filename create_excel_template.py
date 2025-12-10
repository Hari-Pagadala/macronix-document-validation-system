import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Create a new workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Cases"

# Define headers
headers = ["Case Number", "First Name", "Last Name", "Contact Number", "Email", "Address", "State", "District", "Pincode"]

# Add headers to first row
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Sample data
sample_data = [
    ["CASE-001", "Rajesh", "Kumar", "9876543210", "rajesh.kumar@email.com", "123 Main Street Apt 5", "Maharashtra", "Mumbai", "400001"],
    ["CASE-002", "Priya", "Singh", "9876543211", "priya.singh@email.com", "456 Oak Avenue", "Karnataka", "Bangalore", "560001"],
    ["CASE-003", "Amit", "Patel", "9876543212", "amit.patel@email.com", "789 Elm Road", "Gujarat", "Ahmedabad", "380001"],
    ["CASE-004", "Neha", "Sharma", "9876543213", "neha.sharma@email.com", "321 Pine Lane", "Delhi", "Central Delhi", "110001"],
    ["CASE-005", "Vikram", "Desai", "9876543214", "vikram.desai@email.com", "654 Maple Drive", "Telangana", "Hyderabad", "500001"],
    ["CASE-006", "Anjali", "Gupta", "9876543215", "anjali.gupta@email.com", "987 Cedar Street", "Uttar Pradesh", "Lucknow", "226001"],
    ["CASE-007", "Rohan", "Nair", "9876543216", "rohan.nair@email.com", "147 Birch Road", "Kerala", "Kochi", "682001"],
    ["CASE-008", "Divya", "Menon", "9876543217", "divya.menon@email.com", "258 Spruce Avenue", "Tamil Nadu", "Chennai", "600001"],
    ["CASE-009", "Sanjay", "Reddy", "9876543218", "sanjay.reddy@email.com", "369 Walnut Lane", "Andhra Pradesh", "Visakhapatnam", "530001"],
    ["CASE-010", "Pooja", "Verma", "9876543219", "pooja.verma@email.com", "741 Ash Drive", "Rajasthan", "Jaipur", "302001"],
]

# Add sample data
for row_num, row_data in enumerate(sample_data, 2):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.alignment = Alignment(horizontal="left", vertical="center")

# Set column widths
column_widths = [15, 12, 12, 15, 25, 30, 15, 15, 12]
for col_num, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

# Save the workbook
output_path = "SAMPLE_CASES.xlsx"
wb.save(output_path)
print(f"✅ Excel file created successfully: {output_path}")
print(f"📊 Sample data includes 10 test cases (CASE-001 to CASE-010)")
print(f"📌 Ready to upload to the system!")
